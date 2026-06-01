"""Spreadsheet low-cost profiling."""

from __future__ import annotations

import asyncio
import csv
import io
from pathlib import Path
from typing import Any, Dict, List

from app.utils.async_file_utils import async_read_text


class SpreadsheetProfiler:
    @staticmethod
    async def profile(path: Path) -> Dict[str, Any]:
        if path.suffix.lower() == ".csv":
            return await SpreadsheetProfiler._profile_csv(path)
        return await asyncio.to_thread(SpreadsheetProfiler._profile_excel, path)

    @staticmethod
    async def _profile_csv(path: Path) -> Dict[str, Any]:
        rows: List[list[str]] = []
        content = await async_read_text(path, encoding="utf-8-sig")
        reader = csv.reader(io.StringIO(content, newline=""))
        for _, row in zip(range(6), reader):
            rows.append(row)
        return {
            "sheets": [{"name": path.stem, "sample_rows": rows, "row_count": None, "column_count": len(rows[0]) if rows else 0}],
            "sheet_count": 1,
        }

    @staticmethod
    def _profile_excel(path: Path) -> Dict[str, Any]:
        if path.suffix.lower() == ".xls":
            return SpreadsheetProfiler._profile_excel_with_pandas(path)
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            return {"error": f"openpyxl unavailable: {exc}", "sheets": [], "sheet_count": 0}

        try:
            workbook = load_workbook(str(path), read_only=True, data_only=True)
            sheets = []
            for sheet in workbook.worksheets:
                sample_rows = []
                for _, row in zip(range(6), sheet.iter_rows(values_only=True)):
                    sample_rows.append(["" if value is None else str(value) for value in row])
                sheets.append({
                    "name": sheet.title,
                    "row_count": sheet.max_row,
                    "column_count": sheet.max_column,
                    "sample_rows": sample_rows,
                })
            workbook.close()
            return {"sheets": sheets, "sheet_count": len(sheets)}
        except Exception as exc:
            return {"error": f"openpyxl failed: {exc}", "sheets": [], "sheet_count": 0}

    @staticmethod
    def _profile_excel_with_pandas(path: Path) -> Dict[str, Any]:
        try:
            import pandas as pd

            excel = pd.ExcelFile(path)
            sheets = []
            for sheet_name in excel.sheet_names:
                frame = pd.read_excel(excel, sheet_name=sheet_name, nrows=6, header=None)
                sample_rows = [
                    ["" if pd.isna(value) else str(value) for value in row]
                    for row in frame.values.tolist()
                ]
                sheets.append({
                    "name": str(sheet_name),
                    "row_count": None,
                    "column_count": len(sample_rows[0]) if sample_rows else 0,
                    "sample_rows": sample_rows,
                })
            return {"sheets": sheets, "sheet_count": len(sheets)}
        except Exception as exc:
            return {"error": f"pandas failed: {exc}", "sheets": [], "sheet_count": 0}
